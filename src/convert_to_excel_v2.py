import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_NUMBER_00
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from ReportEmailer import ReportEmailer, SMTPConfig
from datetime import date, timedelta

# --- Config ---
# Point this to your enriched CSV. The code will also work with the original CSV.
INPUT_CSV = "assets/examples_and_output/all_data.csv"
OUTPUT_XLSX = "assets/wh_sales_cases_by_warehouse.xlsx"

# For the report file name
def last_weeks_sunday() -> str:
    today = date.today()
    # weekday(): Monday=0, Sunday=6
    days_since_sunday = (today.weekday() + 1) % 7
    # Go back to this week's Sunday, then subtract 7 days
    last_sunday = today - timedelta(days=days_since_sunday + 7)
    return last_sunday.strftime("%Y-%m-%d")

# Load CSV file (auto-detect delimiter; works for CSV or TSV)
# If you KNOW it's comma-delimited, you can switch back to: pd.read_csv(INPUT_CSV)
df = pd.read_csv(INPUT_CSV, sep=None, engine='python')

# Rename columns for clarity (keep original mappings; add enriched ones if present)
rename_map = {
    'week_start': 'Week Start',
    'warehouse': 'Warehouse',
    'sales': 'Sales ($)',
    'cases': 'Cases',
    'cost_per_case': 'Sales/Case ($)',
    'raw_labor_cost/case_goal': 'Raw Labor Cost/Case Goal ($)',
    'labor_cost_with_pto/case_goal': 'Labor Cost w/ PTO/Case Goal ($)',
    'loaded_labor_cost/case_goal': 'Loaded Labor Cost/Case Goal ($)',
    # Enriched columns (only applied if present)
    'raw_labor_cost': 'Raw Labor Cost ($)',
    'raw_labor_hours': 'Raw Labor Hours',
    'cases/hr': 'Cases/Hr',
    'raw_labor_cost/case': 'Raw Labor Cost/Case ($)',
    'labor_cost_with_pto': 'Labor Cost w/ PTO ($)',
    'labor_cost_with_pto/case': 'Labor Cost w/ PTO/Case ($)',
    'loaded_labor_cost': 'Loaded Labor Cost ($)',
    'loaded_labor_cost/case': 'Loaded Labor Cost/Case ($)'
}
df.rename(columns=rename_map, inplace=True)

# Normalize warehouse SP to HA
if 'Warehouse' in df.columns:
    mask_sp = df['Warehouse'].astype(str).str.strip().eq('SP')
    df.loc[mask_sp, 'Warehouse'] = 'HA'

preferred_col_order = ['Week Start', 'Warehouse', 'Cases', 'Sales ($)', 'Sales/Case ($)',
                       'Raw Labor Hours',       'Cases/Hr',
                       'Raw Labor Cost ($)',    'Raw Labor Cost/Case ($)',      'Raw Labor Cost/Case Goal ($)',
                       'Labor Cost w/ PTO ($)', 'Labor Cost w/ PTO/Case ($)',   'Labor Cost w/ PTO/Case Goal ($)',
                       'Loaded Labor Cost ($)', 'Loaded Labor Cost/Case ($)',   'Loaded Labor Cost/Case Goal ($)']

# Ensure numeric types for formatting (only convert columns that exist)
numeric_columns = [
    'Sales ($)', 'Cases', 'Sales/Case ($)',
    'Raw Labor Cost ($)', 'Raw Labor Hours', 'Cases/Hr',
    'Raw Labor Cost/Case ($)', 'Labor Cost w/ PTO ($)', 'Labor Cost w/ PTO/Case ($)',
    'Loaded Labor Cost ($)', 'Loaded Labor Cost/Case ($)', 'Raw Labor Cost/Case Goal ($)',
    'Labor Cost w/ PTO/Case Goal ($)', 'Loaded Labor Cost/Case Goal ($)'
]
for col in numeric_columns:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

# --- Recompute TOTAL goal columns via cases-weighted average ---
# Will process all three goal columns uniformly.
goal_headers = [
    'Raw Labor Cost/Case Goal ($)',
    'Labor Cost w/ PTO/Case Goal ($)',
    'Loaded Labor Cost/Case Goal ($)',
]

# Normalize Warehouse just for filtering (TOTAL detection)
wh = df['Warehouse'].astype(str).str.strip().str.upper()
non_total_mask = wh != 'TOTAL'

# Denominator by week: TOTAL Cases per week (used for every goal column)
total_cases_by_week = (
    df.loc[wh == 'TOTAL']
      .set_index('Week Start')['Cases']
)

for goal_col in goal_headers:
    if goal_col not in df.columns:
        continue  # skip gracefully if a goal column is missing

    # Weighted numerator by week: sum(Cases * Goal) over non-TOTAL warehouses
    weighted = df.loc[non_total_mask, 'Cases'] * df.loc[non_total_mask, goal_col]
    numerator_by_week = weighted.groupby(df.loc[non_total_mask, 'Week Start']).sum()

    # Align denominators to numerator index; avoid divide-by-zero (or missing TOTAL cases)
    denom = total_cases_by_week.reindex(numerator_by_week.index)
    safe_denom = denom.where(denom.ne(0))
    total_goal_by_week = numerator_by_week.div(safe_denom)

    # Overwrite TOTAL rows’ values for this goal column
    total_mask = (wh == 'TOTAL')
    df.loc[total_mask, goal_col] = df.loc[total_mask, 'Week Start'].map(total_goal_by_week)
# --- End recompute TOTAL goal columns ---

# Create a new workbook
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define alternating fill colors
fill_grey = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
fill_blue = PatternFill(start_color="B4E2F1", end_color="B4E2F1", fill_type="solid")

# Function to apply formatting to columns
def apply_column_formatting(ws, headers, header_row=1):
    # Define separate sets for formatting
    currency_headers = {
        'Sales ($)', 'Sales/Case ($)',  # keep Sales/Case at currency default (2 decimals)
        'Raw Labor Cost ($)', 'Labor Cost w/ PTO ($)', 'Loaded Labor Cost ($)'
    }
    # REMOVE the four_decimal_headers set entirely
    comma_headers = {'Cases', 'Raw Labor Hours'}
    decimal_headers = {'Cases/Hr'}
    for col_idx, header in enumerate(headers, start=1):
        if header in currency_headers:
            fmt = FORMAT_CURRENCY_USD_SIMPLE # '"$"#,##0.00_-'
        elif header in comma_headers:
            fmt = '#,##0'  # Comma-separated integers
        elif header in decimal_headers:
            fmt = FORMAT_NUMBER_00  # Two decimals
        else:
            fmt = None

        if fmt:
            for row in ws.iter_rows(min_row=header_row+1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = fmt

def apply_precision_by_sheet(ws, headers, is_total_sheet: bool, header_row: int = 1):
    """
    Set number formats for per-case metrics per worksheet:
      - Non-TOTAL sheets:
          Actual per-case -> 3 decimals
          Goal  per-case -> 2 decimals
      - TOTAL sheet:
          Both actual & goal per-case -> 4 decimals
    """
    # Target columns (by header text)
    actual_per_case_cols = {
        'Raw Labor Cost/Case ($)',
        'Labor Cost w/ PTO/Case ($)',
        'Loaded Labor Cost/Case ($)',
    }
    goal_per_case_cols = {
        'Raw Labor Cost/Case Goal ($)',
        'Labor Cost w/ PTO/Case Goal ($)',
        'Loaded Labor Cost/Case Goal ($)',
    }

    # Choose format strings (currency with thousands and specified decimals)
    if is_total_sheet:
        fmt_actual = '"$"#,##0.0000'
        fmt_goal   = '"$"#,##0.0000'
    else:
        fmt_actual = '"$"#,##0.000'
        fmt_goal   = '"$"#,##0.00'

    last_row = ws.max_row
    data_start = header_row + 1

    def _apply_format_to_header_set(header_set, fmt):
        for h in header_set:
            if h in headers:
                col_idx = headers.index(h) + 1
                for row in ws.iter_rows(min_row=data_start, max_row=last_row,
                                        min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = fmt

    # Apply formats to both sets
    _apply_format_to_header_set(actual_per_case_cols, fmt_actual)
    _apply_format_to_header_set(goal_per_case_cols,   fmt_goal)

def apply_goal_coloring(ws, headers, header_row=1):
    """
    Color actuals red if actual > goal, green if actual <= goal.
    Applies conditional formatting per column for all data rows.
    """
    pairs = [
        ("Raw Labor Cost/Case ($)", "Raw Labor Cost/Case Goal ($)"),
        ("Labor Cost w/ PTO/Case ($)", "Labor Cost w/ PTO/Case Goal ($)"),
        ("Loaded Labor Cost/Case ($)", "Loaded Labor Cost/Case Goal ($)")
    ]

    last_row = ws.max_row
    data_start = header_row + 1

    for actual, goal in pairs:
        if actual in headers and goal in headers:
            a_idx = headers.index(actual) + 1
            g_idx = headers.index(goal) + 1

            a_col = get_column_letter(a_idx)
            g_col = get_column_letter(g_idx)

            # Target the actual column's data range
            cell_range = f"{a_col}{data_start}:{a_col}{last_row}"

            # Red text when actual > goal
            red_rule = FormulaRule(
                formula=[f"{a_col}{data_start}>{g_col}{data_start}"],
                font=Font(color="FF0000")  # red
            )
            ws.conditional_formatting.add(cell_range, red_rule)

            # Green text when actual <= goal
            green_rule = FormulaRule(
                formula=[f"{a_col}{data_start}<={g_col}{data_start}"],
                font=Font(color="006100")  # dark green
            )
            ws.conditional_formatting.add(cell_range, green_rule)

def color_goal_columns_yellow(ws, headers, header_row=1, include_header=False):
    """
    Color the three goal columns' font yellow.
    By default, only data rows are colored (not the header).
    """
    goal_headers = [
        "Raw Labor Cost/Case Goal ($)",
        "Labor Cost w/ PTO/Case Goal ($)",
        "Loaded Labor Cost/Case Goal ($)",
    ]

    start_row = header_row if include_header else header_row + 1
    last_row = ws.max_row

    for header in goal_headers:
        if header in headers:
            col_idx = headers.index(header) + 1  # 1-based
            for row in ws.iter_rows(min_row=start_row, max_row=last_row,
                                    min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.font = Font(color="CAAF18")  # yellow text

# --- Per-warehouse sheets ---
if 'Warehouse' not in df.columns:
    raise ValueError("Expected 'Warehouse' column not found after renaming. Check input file and headers.")

for wh, group in df.groupby('Warehouse', sort=False):
    sheet_name = "HA" if str(wh) == "SP" else str(wh)
    ws = wb.create_sheet(title=sheet_name[:31])

    # ----- Build headers in preferred order -----
    INCLUDE_WAREHOUSE_COLUMN = True

    group_cols = list(group.columns)
    if not INCLUDE_WAREHOUSE_COLUMN and 'Warehouse' in group_cols:
        group_cols.remove('Warehouse')

    ordered_first = [c for c in preferred_col_order if c in group_cols]
    remaining = [c for c in group_cols if c not in ordered_first]
    headers = ordered_first + remaining
    # -------------------------------------------

    # Write header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # --- Freeze the header row ---
    ws.freeze_panes = 'A2'

    # --- Center-align the header row ---
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows with alternating fills (robust: use current last row, skip fully-empty)
    for _, row in group.iterrows():
        values = [row.get(col) for col in headers]
        # Guard: skip rows that are entirely empty/NaN (prevents blank lines if present in source)
        if not any(pd.notna(v) and v != '' for v in values):
            continue

        ws.append(values)
        r = ws.max_row  # the row we just appended

        # Alternate fill based on *data* row index (exclude header row 1)
        fill = fill_grey if ((r - 1) % 2 == 0) else fill_blue
        for cell in ws[r]:
            cell.fill = fill

    # ---------- NEW: Append a YTD totals row at the bottom ----------
    def safe_sum(col):
        return group[col].sum(skipna=True) if col in group.columns else np.nan

    def safe_div(a, b):
        return (a / b) if (pd.notna(a) and pd.notna(b) and b != 0) else np.nan

    def weighted_goal(col):
        if col not in group.columns:
            return np.nan
        valid = group[['Cases', col]].dropna()
        if valid.empty:
            return np.nan
        wsum = (valid['Cases'] * valid[col]).sum()
        denom = valid['Cases'].sum()
        return (wsum / denom) if denom != 0 else np.nan

    totals = {}

    # Label + identity
    totals['Week Start'] = 'YTD'
    if 'Warehouse' in group.columns:
        totals['Warehouse'] = str(wh)

    # Additive sums
    totals['Cases'] = safe_sum('Cases')
    totals['Sales ($)'] = safe_sum('Sales ($)')
    totals['Raw Labor Hours'] = safe_sum('Raw Labor Hours')
    totals['Raw Labor Cost ($)'] = safe_sum('Raw Labor Cost ($)')
    totals['Labor Cost w/ PTO ($)'] = safe_sum('Labor Cost w/ PTO ($)')
    totals['Loaded Labor Cost ($)'] = safe_sum('Loaded Labor Cost ($)')

    # Recomputed rates/ratios
    totals['Sales/Case ($)'] = safe_div(totals['Sales ($)'], totals['Cases'])
    totals['Raw Labor Cost/Case ($)'] = safe_div(totals['Raw Labor Cost ($)'], totals['Cases'])
    totals['Labor Cost w/ PTO/Case ($)'] = safe_div(totals['Labor Cost w/ PTO ($)'], totals['Cases'])
    totals['Loaded Labor Cost/Case ($)'] = safe_div(totals['Loaded Labor Cost ($)'], totals['Cases'])
    totals['Cases/Hr'] = safe_div(totals['Cases'], totals['Raw Labor Hours'])

    # Cases-weighted goals across all rows in this sheet
    totals['Raw Labor Cost/Case Goal ($)'] = weighted_goal('Raw Labor Cost/Case Goal ($)')
    totals['Labor Cost w/ PTO/Case Goal ($)'] = weighted_goal('Labor Cost w/ PTO/Case Goal ($)')
    totals['Loaded Labor Cost/Case Goal ($)'] = weighted_goal('Loaded Labor Cost/Case Goal ($)')

    # Append totals aligned to visible headers
    total_values = [totals.get(col, None) for col in headers]
    ws.append(total_values)
    total_row_idx = ws.max_row

    # Apply a black border (box) around the totals row and make it bold
    thick = Side(style='thick', color='000000')
    min_col = 1
    max_col = len(headers)
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=thick,
            bottom=thick,
            left=thick if col_idx == min_col else Side(style=None),
            right=thick if col_idx == max_col else Side(style=None),
        )
    # ---------- End YTD totals row ----------

    # Apply base formats, then per-sheet precision overrides
    apply_column_formatting(ws, headers)
    apply_precision_by_sheet(ws, headers, is_total_sheet=(sheet_name.strip().upper() == 'TOTAL'))

    # Apply conditional coloring for goals (includes totals row; say the word if you want it excluded)
    apply_goal_coloring(ws, headers, header_row=1)

    # Color goal columns' font yellow
    color_goal_columns_yellow(ws, headers, header_row=1)

    # Autosize columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length * 1.35

# --- Save and Email ---
# Save
wb.save(OUTPUT_XLSX)
"""
# Email
smtp = SMTPConfig(
    host="smtp.office365.com",  # Example for Exchange Online via SMTP AUTH
    port=587,
    username="jsommerfeldt@russdaviswholesale.com",
    password="Davis@2025",
    use_starttls=True,
    use_ssl=False,
    sender="jsommerfeldt@russdaviswholesale.com",
)

emailer = ReportEmailer(smtp)
emailer.send_xlsx_report(
    data_source=OUTPUT_XLSX,
    recipients=["BHopkins@russdaviswholesale.com", "jgraybill@russdaviswholesale.com", "BFihn@russdaviswholesale.com",
                "kmoran@russdaviswholesale.com", "jsommerfeldt@russdaviswholesale.com"],
    subject=f"[Automated Email] - Branch Labor {last_weeks_sunday()}",
    body_text=f"This is an automated email containing the Branch Labor report, with {last_weeks_sunday()}'s data included.\n\nPlease reach out to Jake if there are any issues with the report.",
    sheet_name=f"Summary",
    index=False,
    attachment_filename=f"Branch_Labor_{last_weeks_sunday()}.xlsx",
    cc=None,
    bcc=None,
    reply_to=None,
)
"""