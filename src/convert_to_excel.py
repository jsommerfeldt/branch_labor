import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_NUMBER_00

# Load CSV file
df = pd.read_csv("assets/wh_sales_cases.csv")

# Rename columns for clarity
df.rename(columns={
    'week_start': 'Week Start',
    'warehouse': 'Warehouse',
    'sales': 'Sales ($)',
    'cases': 'Cases',
    'cost_per_case': 'Cost/Case ($)'
}, inplace=True)

# Ensure numeric types for formatting
df['Sales ($)'] = pd.to_numeric(df['Sales ($)'], errors='coerce')
df['Cases'] = pd.to_numeric(df['Cases'], errors='coerce')
df['Cost/Case ($)'] = pd.to_numeric(df['Cost/Case ($)'], errors='coerce')

# Create a new workbook
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define alternating fill colors
fill_grey = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
fill_blue = PatternFill(start_color="B4E2F1", end_color="B4E2F1", fill_type="solid")

# Function to apply formatting to columns
def apply_column_formatting(ws, header_row=1):
    headers = [cell.value for cell in ws[header_row]]
    for col_idx, header in enumerate(headers, start=1):
        if header in ['Sales ($)', 'Cost/Case ($)']:
            for row in ws.iter_rows(min_row=header_row+1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        elif header == 'Cases':
            for row in ws.iter_rows(min_row=header_row+1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = FORMAT_NUMBER_00

# Create individual sheets by warehouse
for wh, group in df.groupby('Warehouse', sort=False):
    ws = wb.create_sheet(title=str(wh)[:31])
    headers = [col for col in group.columns if col != 'Warehouse']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row_idx, (_, row) in enumerate(group.iterrows(), start=2):
        values = [row[col] for col in headers]
        ws.append(values)
        fill = fill_grey if row_idx % 2 == 0 else fill_blue
        for cell in ws[row_idx]:
            cell.fill = fill
    apply_column_formatting(ws)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

# Create Totals sheet
agg = df.groupby('Week Start').agg({
    'Sales ($)': 'sum',
    'Cases': 'sum',
    'Cost/Case ($)': 'mean'
}).reset_index()

ws_totals = wb.create_sheet(title='Totals')
headers = list(agg.columns)
ws_totals.append(headers)
for cell in ws_totals[1]:
    cell.font = Font(bold=True)
for row_idx, row in enumerate(agg.itertuples(index=False), start=2):
    ws_totals.append(list(row))
    fill = fill_grey if row_idx % 2 == 0 else fill_blue
    for cell in ws_totals[row_idx]:
        cell.fill = fill
apply_column_formatting(ws_totals)
for col in ws_totals.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws_totals.column_dimensions[col[0].column_letter].width = max_length + 2

# Save workbook
wb.save("assets/wh_sales_cases_by_warehouse.xlsx")
