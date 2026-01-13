
# aload.py
"""
Create Excel workbooks from aall_data.csv:
- One workbook per acc_year
- One worksheet per warehouse

IMPORTANT:
- This script intentionally uses the complex STYLING approach from convert_to_excel_v2.py
  (header formatting, alternating fills, number formats, conditional formatting, goal font color,
  autosizing, etc.)
- It intentionally does NOT compute / recompute any numeric values.
  If a column is not present in the CSV, it is left blank.

Input:
    assets\\examples_and_output\\aall_data.csv

Output:
    assets\\examples_and_output\\wh_sales_cases_by_warehouse_<acc_year>.xlsx

Dependencies:
    pip install pandas numpy openpyxl
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional
from datetime import date
import json 

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter


# ----------------------------
# Config
# ----------------------------
INPUT_CSV = Path(r"assets\examples_and_output\aall_data.csv")
OUTPUT_DIR = Path(r"assets\examples_and_output")
KPI_GOALS_PATH = Path(r"assets\dict\historical_KPI_goals_sp-ha.json")

# If True, appends a "YTD" row but leaves values blank (styled only).
# (No numeric logic; purely a styled placeholder row.)
APPEND_STYLED_YTD_ROW = True

# If True, appends a styled "TOTALS" row but leaves values blank (styled only).
APPEND_STYLED_TOTAL_ROW = False

ACC_YEAR_START = {
    "2025": date(2024, 12, 29),
    "2026": date(2025, 12, 28),
    "2027": date(2026, 12, 27),
}


# ----------------------------
# Styling (sourced from convert_to_excel_v2.py patterns)
# ----------------------------
fill_grey = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
fill_blue = PatternFill(start_color="B4E2F1", end_color="B4E2F1", fill_type="solid")


def apply_column_formatting(ws, headers, header_row: int = 1) -> None:
    """
    Apply base number formats by header name.
    Mirrors convert_to_excel_v2.py patterns (styling only). [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
    """
    currency_headers = {
        "Sales ($)",
        "Sales/Case ($)",  # currency default (2 decimals)
        "Raw Labor Cost ($)",
        "Labor Cost w/ PTO ($)",
        "Loaded Labor Cost ($)",
    }
    comma_headers = {"Total Cases", "Sale Cases", "Raw Labor Hours"}
    decimal_headers = {"Cases/Hr"}

    for col_idx, header in enumerate(headers, start=1):
        if header in currency_headers:
            fmt = FORMAT_CURRENCY_USD_SIMPLE
        elif header in comma_headers:
            fmt = "#,##0"
        elif header in decimal_headers:
            fmt = FORMAT_NUMBER_00
        else:
            fmt = None

        if fmt:
            for row in ws.iter_rows(
                min_row=header_row + 1,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    cell.number_format = fmt


def apply_precision_by_sheet(ws, headers, is_total_sheet: bool, header_row: int = 1) -> None:
    """
    Per-sheet precision rules for per-case metrics, matching convert_to_excel_v2.py. [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
    - Non-TOTAL: actual per-case 3 decimals; goal per-case 2 decimals
    - TOTAL: both actual+goal per-case 4 decimals
    """
    actual_per_case_cols = {
        "Raw Labor Cost/Case ($)",
        "Labor Cost w/ PTO/Case ($)",
        "Loaded Labor Cost/Case ($)",
    }
    goal_per_case_cols = {
        "Raw Labor Cost/Case Goal ($)",
        "Labor Cost w/ PTO/Case Goal ($)",
        "Loaded Labor Cost/Case Goal ($)",
    }

    if is_total_sheet:
        fmt_actual = '"$"#,##0.0000'
        fmt_goal = '"$"#,##0.0000'
    else:
        fmt_actual = '"$"#,##0.000'
        fmt_goal = '"$"#,##0.00'

    last_row = ws.max_row
    data_start = header_row + 1

    def _apply(header_set, fmt):
        for h in header_set:
            if h in headers:
                col_idx = headers.index(h) + 1
                for row in ws.iter_rows(
                    min_row=data_start,
                    max_row=last_row,
                    min_col=col_idx,
                    max_col=col_idx,
                ):
                    for cell in row:
                        cell.number_format = fmt

    _apply(actual_per_case_cols, fmt_actual)
    _apply(goal_per_case_cols, fmt_goal)


def apply_goal_coloring(ws, headers, header_row: int = 1) -> None:
    """
    Conditional formatting: actual red if actual > goal, green otherwise.
    Mirrors convert_to_excel_v2.py (styling only). [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)

    Note: If goal columns are blank (common with aall_data.csv), the rules remain but
    won't meaningfully color values—still preserves styling behavior.
    """
    pairs = [
        ("Raw Labor Cost/Case ($)", "Raw Labor Cost/Case Goal ($)"),
        ("Labor Cost w/ PTO/Case ($)", "Labor Cost w/ PTO/Case Goal ($)"),
        ("Loaded Labor Cost/Case ($)", "Loaded Labor Cost/Case Goal ($)"),
    ]
    last_row = ws.max_row
    data_start = header_row + 1

    for actual, goal in pairs:
        if actual in headers and goal in headers and last_row >= data_start:
            a_idx = headers.index(actual) + 1
            g_idx = headers.index(goal) + 1
            a_col = get_column_letter(a_idx)
            g_col = get_column_letter(g_idx)

            cell_range = f"{a_col}{data_start}:{a_col}{last_row}"

            # Red when actual > goal
            red_rule = FormulaRule(
                formula=[f"{a_col}{data_start}>{g_col}{data_start}"],
                font=Font(color="FF0000"),
            )
            ws.conditional_formatting.add(cell_range, red_rule)

            # Green when actual <= goal
            green_rule = FormulaRule(
                formula=[f"{a_col}{data_start}<={g_col}{data_start}"],
                font=Font(color="006100"),
            )
            ws.conditional_formatting.add(cell_range, green_rule)


def color_goal_columns_yellow(ws, headers, header_row: int = 1, include_header: bool = False) -> None:
    """
    Yellow font for goal columns (data rows by default), as in convert_to_excel_v2.py. [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
    """
    goal_headers = [
        "Raw Labor Cost/Case Goal ($)",
        "Labor Cost w/ PTO/Case Goal ($)",
        "Loaded Labor Cost/Case Goal ($)",
    ]
    start_row = header_row if include_header else header_row + 1
    last_row = ws.max_row

    for header in goal_headers:
        if header in headers:
            col_idx = headers.index(header) + 1
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=last_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    cell.font = Font(color="CAAF18")  # yellow text


def autosize_columns(ws) -> None:
    """
    Autosize columns using the same simple strategy as convert_to_excel_v2.py. [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
    """
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max(8, max_length * 1.35)


def apply_thick_box_border_row(ws, row_idx: int, n_cols: int) -> None:
    """
    Apply thick black box border around an entire row, and bold font.
    This mimics the totals-row styling pattern in convert_to_excel_v2.py,
    but does NOT compute totals. [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
    """
    thick = Side(style="thick", color="000000")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=thick,
            bottom=thick,
            left=thick if col_idx == 1 else Side(style=None),
            right=thick if col_idx == n_cols else Side(style=None),
        )


# ----------------------------
# Sheet naming
# ----------------------------
_INVALID_SHEET_CHARS = r"[\[\]\*:/\\\?]"
_MAX_SHEET_LEN = 31


def sanitize_sheet_name(name: str) -> str:
    safe = re.sub(_INVALID_SHEET_CHARS, "_", str(name).strip())
    safe = safe[:_MAX_SHEET_LEN] if len(safe) > _MAX_SHEET_LEN else safe
    return safe if safe else "Sheet"


def dedupe_sheet_names(names: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for n in names:
        base = sanitize_sheet_name(n)
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            i = seen[base] + 1
            seen[base] = i
            suffix = f"_{i}"
            trimmed = base[: _MAX_SHEET_LEN - len(suffix)]
            out.append(f"{trimmed}{suffix}")
    return out


# ----------------------------
# Column mapping / layout
# ----------------------------
# Desired header order is taken from convert_to_excel_v2.py's preferred_col_order. [1](https://russdaviswholesale-my.sharepoint.com/personal/jsommerfeldt_russdaviswholesale_com/Documents/Microsoft%20Copilot%20Chat%20Files/convert_to_excel_v2.py)
PREFERRED_HEADERS = [
    "Week Start",
    "Warehouse",
    "Total Cases",
    "Sale Cases",
    "Sales ($)",
    "Raw Labor Hours",
    "Cases/Hr",
    "Raw Labor Cost ($)",
    "Raw Labor Cost/Case ($)",
    "Raw Labor Cost/Case Goal ($)",
    "Labor Cost w/ PTO ($)",
    "Labor Cost w/ PTO/Case ($)",
    "Labor Cost w/ PTO/Case Goal ($)",
    "Loaded Labor Cost ($)",
    "Loaded Labor Cost/Case ($)",
    "Loaded Labor Cost/Case Goal ($)",
]

# Map aall_data.csv columns -> styled header names (no computations)
#   Note: Sales/Case ($) and Goal columns are not in aall_data.csv; left blank.
CSV_TO_HEADER = {
    "week_start": "Week Start",
    "warehouse": "Warehouse",
    "all_cases": "Total Cases",
    "cases": "Sale Cases",
    "sales": "Sales ($)",
    "raw_labor_hours": "Raw Labor Hours",
    "cases/hr": "Cases/Hr",
    "raw_labor_cost": "Raw Labor Cost ($)",
    "raw_labor_cost/case": "Raw Labor Cost/Case ($)",
    "raw_labor_cost/case_goal": "Raw Labor Cost/Case Goal ($)",
    "labor_cost_with_pto": "Labor Cost w/ PTO ($)",
    "labor_cost_with_pto/case": "Labor Cost w/ PTO/Case ($)",
    "labor_cost_with_pto/case_goal": "Labor Cost w/ PTO/Case Goal ($)",
    "loaded_labor_cost": "Loaded Labor Cost ($)",
    "loaded_labor_cost/case": "Loaded Labor Cost/Case ($)",
    "loaded_labor_cost/case_goal": "Loaded Labor Cost/Case Goal ($)",

}


def build_output_headers(df_cols: List[str]) -> List[str]:
    """
    Output headers:
    - Use PREFERRED_HEADERS first
    - Then append any extra mapped headers from CSV not in preferred list
    """
    mapped_headers = [CSV_TO_HEADER[c] for c in df_cols if c in CSV_TO_HEADER]
    ordered_first = [h for h in PREFERRED_HEADERS if h in set(mapped_headers) or h in PREFERRED_HEADERS]
    # Keep the preferred order fixed (even if blank columns)
    # Append non-preferred mapped columns at the end (if any)
    remaining = [h for h in mapped_headers if h not in ordered_first]
    return ordered_first + remaining


def row_values_from_df_row(row: pd.Series, headers: List[str]) -> List[object]:
    """
    Build a row for Excel by matching headers to CSV columns using CSV_TO_HEADER.
    Columns not present are left blank (None). No computations.
    """
    header_to_csv = {v: k for k, v in CSV_TO_HEADER.items()}
    values = []
    for h in headers:
        csv_col = header_to_csv.get(h)
        values.append(row.get(csv_col) if csv_col in row.index else None)
    return values


# ----------------------------
# Main generation
# ----------------------------

# -- Year-specific goal logic (NEW) --
def inject_2025_goals(df_year: pd.DataFrame, acc_year: str, goals_path: Path) -> pd.DataFrame:
    """
    For acc_year=2025: load historical monthly (and year) goals and expand to weekly,
    then left-join to df_year by (week_start, warehouse).
    """
    goals_df = load_cost_per_case_df(goals_path)  # JSON contains 2025 monthly anchors + "2025" row
    goals_df = expand_monthly_kpis_to_weeks(goals_df)

    # Restrict to this accounting year's calendar window
    start = ACC_YEAR_START[acc_year]
    end_excl = ACC_YEAR_START.get(str(int(acc_year) + 1), None)
    if end_excl is None:
        # If next year's start is unknown, include weeks within calendar year 2025
        end_excl = pd.Timestamp(start) + pd.Timedelta(days=366)  # safe upper bound

    weekly_goals = goals_df[goals_df["week_start"].apply(lambda x: not isinstance(x, str))].copy()
    weekly_goals["week_start"] = pd.to_datetime(weekly_goals["week_start"]).dt.date
    mask = (weekly_goals["week_start"] >= start) & (weekly_goals["week_start"] < end_excl)
    weekly_goals = weekly_goals.loc[mask]

    # Keep only the goal columns we know about
    keep_cols = ["week_start", "warehouse",
                 "raw_labor_cost/case_goal",
                 "labor_cost_with_pto/case_goal",
                 "loaded_labor_cost/case_goal"]
    weekly_goals = weekly_goals[keep_cols]

    # Left-join into this year's data by (week_start, warehouse)
    out = df_year.merge(weekly_goals, on=["week_start", "warehouse"], how="left", suffixes=("", ""))
    return out


def inject_2026_goals_from_2025_actuals(df_2026: pd.DataFrame, df_2025: pd.DataFrame) -> pd.DataFrame:
    """
    For acc_year=2026: for each warehouse and accounting week_number,
    set weekly goals to the same week_number's *actual per-case* metrics in 2025.
    """
    # Compute week_number in 2025 and 2026
    df_2025 = df_2025.copy()
    df_2025["week_number"] = df_2025["week_start"].apply(lambda d: compute_week_number("2025", d))
    df_2026 = df_2026.copy()
    df_2026["week_number"] = df_2026["week_start"].apply(lambda d: compute_week_number("2026", d))

    # Build lookup for 2025 actual per-case by (warehouse, week_number)
    actual_cols = ["raw_labor_cost/case", "labor_cost_with_pto/case", "loaded_labor_cost/case"]
    lookup = (
        df_2025[["warehouse", "week_number"] + actual_cols]
        .dropna(subset=["week_number"])
        .drop_duplicates(["warehouse", "week_number"])
        .set_index(["warehouse", "week_number"])
    )

    def map_goal(col_actual: str) -> pd.Series:
        # Align (warehouse, week_number) keys
        keys = list(zip(df_2026["warehouse"].str.upper(), df_2026["week_number"]))

        # For each key, read the row at (warehouse, week_number) and take the column `col_actual`
        out = []
        for (w, n) in keys:
            if (w, n) in lookup.index:
                out.append(lookup.loc[(w, n), col_actual])
            else:
                out.append(np.nan)
        series = pd.Series(out, index=df_2026.index)
        return series

    # Create goal columns in 2026 based on 2025 actuals
    df_2026["raw_labor_cost/case_goal"] = map_goal("raw_labor_cost/case")
    df_2026["labor_cost_with_pto/case_goal"] = map_goal("labor_cost_with_pto/case")
    df_2026["loaded_labor_cost/case_goal"] = map_goal("loaded_labor_cost/case")

    # Clean up
    df_2026 = df_2026.drop(columns=["week_number"])
    return df_2026

def scale_goals_by_sales_ratio(df_year: pd.DataFrame) -> pd.DataFrame:
    """
    Convert weekly per-case GOALs from a sale-case basis to a total-case basis by
    multiplying each week's goal by (Sale Cases / Total Cases).
    Safe-guards against division by zero / missing data.
    """
    df_year = df_year.copy()

    # These are the CSV column names that hold the GOAL values in df_year
    goal_cols = [
        "raw_labor_cost/case_goal",
        "labor_cost_with_pto/case_goal",
        "loaded_labor_cost/case_goal",
    ]

    # Require both Sale Cases ("cases") and Total Cases ("all_cases") to be present
    if not {"cases", "all_cases"}.issubset(df_year.columns):
        return df_year

    # ratio = Sale Cases / Total Cases
    ratio = df_year["cases"] / df_year["all_cases"]
    ratio = ratio.replace([np.inf, -np.inf], np.nan)

    for gc in goal_cols:
        if gc in df_year.columns:
            df_year[gc] = df_year[gc] * ratio

    return df_year

def recompute_total_weekly_goals(df_year: pd.DataFrame) -> pd.DataFrame:
    """
    Recompute weekly goals for TOTAL rows as a cases-weighted average across non-TOTAL warehouses
    for each goal column. Uses 'cases' (Sale Cases) for weighting and TOTAL week's cases as denominator.
    """
    goal_cols = ["raw_labor_cost/case_goal", "labor_cost_with_pto/case_goal", "loaded_labor_cost/case_goal"]
    df_year = df_year.copy()
    if not set(goal_cols).issubset(df_year.columns):
        return df_year  # if goal cols not present, nothing to recompute

    wh_upper = df_year["warehouse"].str.upper()
    non_total_mask = wh_upper != "TOTAL"

    # TOTAL denominator: TOTAL "cases" by week
    total_cases_by_week = (
        df_year.loc[wh_upper == "TOTAL"].set_index("week_start")["cases"]
    )

    for gc in goal_cols:
        # Numerator: sum(Cases * goal) across non-TOTAL warehouses for that week
        weighted = df_year.loc[non_total_mask, "cases"] * df_year.loc[non_total_mask, gc]
        numerator_by_week = weighted.groupby(df_year.loc[non_total_mask, "week_start"]).sum(min_count=1)

        denom = total_cases_by_week.reindex(numerator_by_week.index)
        safe_denom = denom.where(denom.ne(0))  # zero→NaN to avoid div-by-zero
        total_goal_by_week = numerator_by_week.div(safe_denom)

        # Overwrite TOTAL weekly goal values
        is_total = wh_upper == "TOTAL"
        mapper = df_year.loc[is_total, "week_start"].map(total_goal_by_week)
        df_year.loc[is_total, gc] = mapper.values

    return df_year

# -- Goals loading/expansion helpers (NEW) --
def load_cost_per_case_df(payroll_path: Path) -> pd.DataFrame:
    """
    Load labor cost per case metrics from a JSON dictionary.
    Expected JSON shape:
    { "2025-01-05": { "JA": {"raw_labor_cost/case_goal": ..., "labor_cost_with_pto/case_goal": ..., ...}, ... }, ... }
    """
    text = payroll_path.read_text(encoding="utf-8").strip()
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            rows = []
            for week, wh_dict in obj.items():
                w = str(week).strip()

                # Keep 4-digit years (e.g., "2025") as year-only rows
                if w.isdigit() and len(w) == 4:
                    week_value = w  # leave as string
                else:
                    # Try to parse as a real date; skip non-date labels (e.g., "YTD", "MTD", "QTD")
                    parsed = pd.to_datetime(w, errors="coerce")
                    if pd.isna(parsed):
                        continue
                    week_value = parsed

                for wh, metrics in (wh_dict or {}).items():
                    row = {
                        "week_start": week_value,
                        "warehouse": str(wh).strip().upper(),
                    }
                    # Dynamically include all metrics
                    for key, value in metrics.items():
                        try:
                            row[key] = float(value)
                        except (TypeError, ValueError):
                            row[key] = float("nan")
                    rows.append(row)
            df = pd.DataFrame(rows)
        else:
            raise ValueError("Unexpected JSON structure")
    except json.JSONDecodeError:
        raise ValueError("Invalid JSON format")

    # Basic cleanup
    if "week_start" in df.columns:
        # Leave 4-digit year strings as-is; ensure dates are Timestamp/Date
        def _norm_week(x):
            if isinstance(x, str) and x.isdigit() and len(x) == 4:
                return x  # year-only rows stay as string
            return pd.to_datetime(x)  # already parseable timestamps
        df["week_start"] = df["week_start"].apply(_norm_week)

    if "warehouse" in df.columns:
        df["warehouse"] = df["warehouse"].astype(str).str.strip().str.upper()

    # Deduplicate
    if not df.empty:
        df = (
            df.sort_values(["week_start", "warehouse"])
              .drop_duplicates(subset=["week_start", "warehouse"], keep="last")
        )
    return df

def expand_monthly_kpis_to_weeks(kpi_df: pd.DataFrame) -> pd.DataFrame:
    # Keep rows where week_start is a 4-char year string (e.g., "2025") untouched
    year_only_df = kpi_df[kpi_df["week_start"].apply(lambda x: isinstance(x, str) and len(x) == 4)].copy()

    # Rows with real dates to expand
    date_rows_df = kpi_df[~kpi_df["week_start"].apply(lambda x: isinstance(x, str) and len(x) == 4)].copy()
    if date_rows_df.empty:
        return kpi_df.copy()

    # Normalize to date for consistent sorting/comparison
    date_rows_df.loc[:, "week_start"] = pd.to_datetime(date_rows_df["week_start"]).dt.date
    date_rows_df = date_rows_df.sort_values("week_start")

    expanded_rows = []

    # Unique anchor dates to expand from
    unique_dates = sorted(date_rows_df["week_start"].unique())

    for i, start_date in enumerate(unique_dates):
        start_ts = pd.Timestamp(start_date)

        if i + 1 < len(unique_dates):
            # End at the next anchor (exclusive)
            end_ts = pd.Timestamp(unique_dates[i + 1])
        else:
            # Calendar-aware: for the final block, end at the first Sunday of the month AFTER next (exclusive)
            month_after_next = start_ts + pd.offsets.MonthBegin(2)
            # First Sunday on/after that date
            offset_days = (6 - month_after_next.weekday()) % 7  # Monday=0 ... Sunday=6
            end_ts = month_after_next + pd.Timedelta(days=offset_days)

        # All rows that share this start_date (each will be replicated weekly)
        current_block = date_rows_df[date_rows_df["week_start"] == start_date]

        # Generate weekly starts up to but not including end_ts (exclusive)
        week = start_ts
        while week < end_ts:
            for _, row in current_block.iterrows():
                new_row = row.copy()
                new_row["week_start"] = week.date()
                expanded_rows.append(new_row)
            week += pd.Timedelta(days=7)

    expanded_df = pd.DataFrame(expanded_rows)

    # Combine expanded weekly rows with untouched year-only rows
    final_df = pd.concat([expanded_df, year_only_df], ignore_index=True)
    return final_df

def compute_week_number(acc_year: str, week_start: object) -> int:
    """
    week_start must be a date (not datetime). Computes accounting week number based on ACC_YEAR_START.
    """
    if acc_year not in ACC_YEAR_START:
        raise ValueError(f"Missing ACC_YEAR_START entry for acc_year={acc_year}")

    if pd.isna(week_start):
        raise ValueError("week_start is NaT/NaN")

    start = ACC_YEAR_START[acc_year]
    delta_days = (week_start - start).days

    if delta_days < 0:
        raise ValueError(f"week_start {week_start} is before acc_year start {start} for acc_year={acc_year}")

    if delta_days % 7 != 0:
        raise ValueError(
            f"week_start {week_start} is not aligned to a Sunday week boundary for acc_year={acc_year} "
            f"(delta_days={delta_days})"
        )

    return 1 + (delta_days // 7)

def load_csv() -> pd.DataFrame:
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"Input CSV not found: {INPUT_CSV}")

    df = pd.read_csv(INPUT_CSV)

    required = {"acc_year", "week_start", "warehouse"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV is missing required columns: {sorted(missing)}")

    # Normalize key columns (not numeric logic; just hygiene)
    df["acc_year"] = df["acc_year"].astype(str).str.strip()
    df["warehouse"] = df["warehouse"].astype(str).str.strip()

    # Keep week_start as string or datetime; Excel will accept either.
    # We'll parse to datetime for better Excel date writing (still not numeric computation).
    df["week_start"] = pd.to_datetime(df["week_start"], errors="coerce")
    df["week_start"] = df["week_start"].dt.date

    # Drop empty keys
    df = df[(df["acc_year"] != "") & (df["warehouse"] != "")]
    return df


def write_styled_sheet(ws, group: pd.DataFrame, sheet_name: str) -> None:
    """
    Write one warehouse sheet with complex styling, without computing any numeric values.
    """
    # Determine headers
    headers = build_output_headers(list(group.columns))

    # Write header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Data rows with alternating fills (same approach as convert_to_excel_v2.py) 
    for _, r in group.iterrows():
        values = row_values_from_df_row(r, headers)
        # Skip fully empty rows
        if not any(pd.notna(v) and v != "" for v in values):
            continue
        ws.append(values)

        excel_row_idx = ws.max_row
        # Alternate based on data row index (exclude header row 1)
        fill = fill_grey if ((excel_row_idx - 1) % 2 == 0) else fill_blue
        for cell in ws[excel_row_idx]:
            cell.fill = fill


    # Optional: append a computed YTD row (cases-weighted goal averages) when enabled (NEW)
    if APPEND_STYLED_YTD_ROW:  # NOTE: ensure the variable name exactly matches your config (APPEND_STYLED_YTD_ROW)
        # Compute weighted goals using the data in 'group'
        def weighted_goal(col_csv: str):
            # Filter rows where both cases and goal are present
            valid = group[[ "all_cases", col_csv ]].dropna()
            if valid.empty:
                return np.nan
            denom = valid["all_cases"].sum()
            if denom == 0:
                return np.nan
            return (valid["all_cases"] * valid[col_csv]).sum() / denom

        # Build a YTD row aligned to current headers
        ytd_values = [None] * len(headers)
        # Label columns
        if "Week Start" in headers:
            ytd_values[headers.index("Week Start")] = "YTD"
        if "Warehouse" in headers:
            ytd_values[headers.index("Warehouse")] = sheet_name

        # Fill the three goal columns (cases-weighted averages across this sheet's rows)
        # CSV column names -> header names
        ytd_goals_map = {
            "raw_labor_cost/case_goal": "Raw Labor Cost/Case Goal ($)",
            "labor_cost_with_pto/case_goal": "Labor Cost w/ PTO/Case Goal ($)",
            "loaded_labor_cost/case_goal": "Loaded Labor Cost/Case Goal ($)",
        }
        for csv_col, header_name in ytd_goals_map.items():
            if header_name in headers:
                val = weighted_goal(csv_col)
                ytd_values[headers.index(header_name)] = val

        ws.append(ytd_values)
        # Style the YTD row (thick border + bold)
        apply_thick_box_border_row(ws, ws.max_row, len(headers))

    if "Week Start" in headers:
        col_idx = headers.index("Week Start") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = "yyyy-mm-dd"

    # Apply formatting rules (styling-only)
    apply_column_formatting(ws, headers)
    apply_precision_by_sheet(ws, headers, is_total_sheet=(sheet_name.strip().upper() == "TOTAL"))
    apply_goal_coloring(ws, headers)
    color_goal_columns_yellow(ws, headers)

    # Autosize columns
    autosize_columns(ws)


def build_workbooks(df: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Pre-slice for cross-year references (e.g., 2026 goals from 2025 actuals)
    df_all = df.copy()
    df_2025_all = df_all[df_all["acc_year"].astype(str) == "2025"].copy()

    for acc_year, df_year in df.groupby("acc_year", dropna=False):
        acc_year_str = str(acc_year).strip()
        if not acc_year_str:
            continue

        # --- Inject goals per year ---
        if acc_year_str == "2025":
            df_year = inject_2025_goals(df_year, acc_year_str, KPI_GOALS_PATH)
        elif acc_year_str == "2026":
            df_year = inject_2026_goals_from_2025_actuals(df_year, df_2025_all)

        # --- Recompute TOTAL weekly goals (cases-weighted average across warehouses) ---
        df_year = recompute_total_weekly_goals(df_year)

        # --- NEW: Convert weekly goals to a total-case basis for 2025 ---
        if acc_year_str == "2025":
            df_year = scale_goals_by_sales_ratio(df_year)

        # Continue with your existing workbook creation
        out_path = OUTPUT_DIR / f"wh_sales_cases_by_warehouse_{acc_year_str}.xlsx"
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Normalize warehouse sheet names (SP->HA) for tab naming only
        df_year = df_year.copy()

        # One sheet per warehouse
        warehouses = sorted(df_year["warehouse"].unique().tolist())
        sheet_names = dedupe_sheet_names(warehouses)
        wh_to_sheet = dict(zip(warehouses, sheet_names))

        for wh in warehouses:
            sheet_name = wh_to_sheet[wh]
            ws = wb.create_sheet(title=sheet_name[:31])

            # Select rows for this warehouse; sort by week_start ascending
            group = df_year[df_year["warehouse"] == wh].copy()
            group = group.sort_values("week_start", ascending=True)

            write_styled_sheet(ws, group, sheet_name=sheet_name)

        wb.save(out_path)
        print(f"Saved: {out_path}")


def main() -> None:
    df = load_csv()
    build_workbooks(df)


if __name__ == "__main__":
    main()
